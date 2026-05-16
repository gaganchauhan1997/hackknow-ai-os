"""
Excel skill — full pipeline.

Capabilities:
  - analyse_csv  : read CSV/XLSX, run pandas describe, save analysis as JSON.
  - build_dashboard : produce a Plotly HTML dashboard from a CSV/XLSX.
  - build_workbook : produce a formatted .xlsx with multiple sheets, KPI cells,
                     conditional formatting, and a charts sheet.
"""

from __future__ import annotations

import asyncio
import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from config.settings import ROOT
from core.logger import get_logger

log = get_logger("skill:excel")

manifest = {
    "description": "Excel + dashboard generator. action='analyse_csv', 'build_dashboard', 'build_workbook'.",
    "parameters": {
        "type": "object",
        "properties": {
            "action": {"type": "string", "enum": ["analyse_csv", "build_dashboard", "build_workbook"]},
            "input_path": {"type": "string"},
            "output_path": {"type": "string"},
            "title": {"type": "string"},
            "kpi_columns": {"type": "array", "items": {"type": "string"}},
            "chart_types": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["action"],
    },
}

OUT = ROOT / ".cache" / "excel_out"
OUT.mkdir(parents=True, exist_ok=True)


def _load(path: str) -> pd.DataFrame:
    p = Path(path)
    if p.suffix.lower() == ".csv":
        return pd.read_csv(p)
    return pd.read_excel(p)


def _analyse(path: str) -> dict:
    df = _load(path)
    return {
        "rows": len(df),
        "columns": list(df.columns),
        "describe": json.loads(df.describe(include="all").to_json()),
        "null_counts": df.isna().sum().to_dict(),
        "dtypes": {c: str(t) for c, t in df.dtypes.items()},
    }


def _dashboard(path: str, output_path: str | None, title: str) -> dict:
    import plotly.express as px
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    df = _load(path)
    output_path = output_path or str(OUT / f"dashboard_{Path(path).stem}.html")
    numeric = df.select_dtypes(include="number")
    categorical = [c for c in df.columns if c not in numeric.columns]
    cols = numeric.columns[:4]
    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=[f"{c} distribution" for c in cols] or ["—"],
    )
    for i, c in enumerate(cols):
        r, k = (i // 2) + 1, (i % 2) + 1
        fig.add_trace(go.Histogram(x=df[c].dropna(), name=c, nbinsx=30), row=r, col=k)
    fig.update_layout(title=title or "HackKnow Dashboard", template="plotly_dark", height=720)
    fig.write_html(output_path, include_plotlyjs="cdn")
    return {"path": output_path, "rows": len(df), "numeric_cols": list(cols),
            "categorical_cols": categorical[:8]}


def _workbook(path: str, output_path: str | None, title: str,
              kpi_columns: list[str] | None) -> dict:
    df = _load(path)
    output_path = output_path or str(OUT / f"{Path(path).stem}_dashboard.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    # Header style
    header_fill = PatternFill("solid", fgColor="0f1320")
    header_font = Font(bold=True, color="21f0c5")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # KPI sheet
    kpi_ws = wb.create_sheet("KPIs")
    kpi_ws["A1"] = title or "HackKnow KPIs"
    kpi_ws["A1"].font = Font(size=18, bold=True)
    numeric = df.select_dtypes(include="number")
    kpis = (kpi_columns or list(numeric.columns)[:6])
    for i, col in enumerate(kpis):
        kpi_ws.cell(row=3 + i, column=1, value=col).font = Font(bold=True)
        if col in numeric.columns:
            kpi_ws.cell(row=3 + i, column=2, value=float(numeric[col].sum()))
            kpi_ws.cell(row=3 + i, column=3, value=float(numeric[col].mean()))
            kpi_ws.cell(row=3 + i, column=4, value=float(numeric[col].max()))
    kpi_ws["B2"] = "Sum"; kpi_ws["C2"] = "Avg"; kpi_ws["D2"] = "Max"

    # Chart sheet
    if not numeric.empty:
        chart_ws = wb.create_sheet("Charts")
        col_idx = list(df.columns).index(numeric.columns[0]) + 1
        chart = BarChart()
        chart.title = numeric.columns[0]
        data_ref = Reference(ws, min_col=col_idx, min_row=1,
                             max_row=min(50, len(df) + 1), max_col=col_idx)
        chart.add_data(data_ref, titles_from_data=True)
        chart_ws.add_chart(chart, "B2")

    # Conditional formatting on first numeric column
    if numeric.shape[1] > 0:
        col_letter = ws.cell(row=1, column=list(df.columns).index(numeric.columns[0]) + 1).column_letter
        rng = f"{col_letter}2:{col_letter}{len(df) + 1}"
        rule = ColorScaleRule(start_type="min", start_color="0f1320",
                              mid_type="percentile", mid_value=50, mid_color="21f0c5",
                              end_type="max", end_color="fdd33b")
        ws.conditional_formatting.add(rng, rule)

    wb.save(output_path)
    return {"path": output_path, "rows": len(df), "kpis": kpis}


async def run(action: str, **kwargs: Any) -> dict:
    if action == "analyse_csv":
        return await asyncio.to_thread(_analyse, kwargs["input_path"])
    if action == "build_dashboard":
        return await asyncio.to_thread(_dashboard, kwargs["input_path"],
                                       kwargs.get("output_path"),
                                       kwargs.get("title", ""))
    if action == "build_workbook":
        return await asyncio.to_thread(_workbook, kwargs["input_path"],
                                       kwargs.get("output_path"),
                                       kwargs.get("title", ""),
                                       kwargs.get("kpi_columns"))
    raise ValueError(f"unknown excel action: {action}")
